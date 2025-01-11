import os
import sys
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.by import By
from time import sleep
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment

# Initialize the WebDriver and open the Deribit page
exchange_name = 'Deribit'
driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()))
driver.maximize_window()
driver.get("https://www.deribit.com/options/BTC/BTC-10JAN25")
driver.implicitly_wait(5)

# Click the "All" button if available
try:
    button = driver.find_element(By.XPATH, "//button[@type='button' and @value='All']")
    button.click()
    print("Button clicked successfully!")
    driver.implicitly_wait(5)
except Exception as e:
    print("Error finding or clicking the button:", e)

# Fetch page content and save it
title = driver.title
print("Page Title:", title)
html = driver.page_source
with open("page_source.html", "w", encoding="utf-8") as file:
    file.write(html)
print("HTML page saved as 'page_source.html'")

# Parse the page with BeautifulSoup
soup = BeautifulSoup(html, 'html.parser')

# Initialize data storage
scraped_data = {'general': [], 'calls': [], 'puts': [], 'strike': []}

# Extract currency price
svg_element = soup.find("svg", {"currency": "btc"})
price_element = svg_element.find_next_sibling("div")
currency_price = price_element.text if price_element else "Price not found"

# Extract option data (Calls, Puts, and Strike)
base_dives = soup.find_all('div', {'style': 'flex: 1 1 0px;'})  # calls & puts & date
for div in base_dives:
    res = div.find_all('div', {'data-id': True})
    for row in res:
        data_id = row['data-id']
        currency_name = data_id.split('-')[0]
        date = data_id.split('-')[1]
        if len(scraped_data['general']) <= len(scraped_data['calls']):
            scraped_data['general'].append([exchange_name, currency_name, currency_price, date, ''])
        if data_id.endswith('C'):  # for Calls
            size_bid = row.find('div', {'data-colid': 'best_bid_amount'}).text.strip() if row.find('div', {'data-colid': 'best_bid_amount'}) else None
            bid = row.find('div', {'data-colid': 'best_bid_price'}).find_all('span')[1].text.strip() if row.find('div', {'data-colid': 'best_bid_price'}) else None
            mark_price = row.find('div', {'data-colid': 'mark_price'}).find_all('span')[0].text.strip() if row.find('div', {'data-colid': 'mark_price'}) else None
            ask = row.find('div', {'data-colid': 'best_ask_price'}).find_all('span')[1].text.strip() if row.find('div', {'data-colid': 'best_ask_price'}) else None
            size_ask = row.find('div', {'data-colid': 'best_ask_amount'}).text.strip() if row.find('div', {'data-colid': 'best_ask_amount'}) else None
            scraped_data['calls'].append(['Calls', size_bid, bid, mark_price, ask, size_ask])
        elif data_id.endswith('P'):  # for Puts
            size_bid = row.find('div', {'data-colid': 'best_bid_amount'}).text.strip() if row.find('div', {'data-colid': 'best_bid_amount'}) else None
            bid = row.find('div', {'data-colid': 'best_bid_price'}).find_all('span')[1].text.strip() if row.find('div', {'data-colid': 'best_bid_price'}) else None
            mark_price = row.find('div', {'data-colid': 'mark_price'}).find_all('span')[0].text.strip() if row.find('div', {'data-colid': 'mark_price'}) else None
            ask = row.find('div', {'data-colid': 'best_ask_price'}).find_all('span')[1].text.strip() if row.find('div', {'data-colid': 'best_ask_price'}) else None
            size_ask = row.find('div', {'data-colid': 'best_ask_amount'}).text.strip() if row.find('div', {'data-colid': 'best_ask_amount'}) else '-'
            scraped_data['puts'].append(['Puts', size_bid, bid, mark_price, ask, size_ask])

# Extract strike prices
base_dives = soup.find('div', {'style': 'width: 60px; flex: 0 0 60px;'}).find_all('span')  # strike
for row in base_dives:
    scraped_data['strike'].append([row.text.strip()])

driver.quit()

# Save data to Excel
file_name = "Exchange Data 2.xlsx"
workbook = Workbook()
worksheet = workbook.active

# Prepare headers and data
data_list = [['Exchange', 'Crypto', 'Price (Spot)', 'Date', 'Days', 'Strike', 'Calls', 'Bid Size', 'Bid', 'Mark Price', 'Ask', 'Ask Size',
              'Puts', 'Bid Size', 'Bid', 'Mark Price', 'Ask', 'Ask Size']]
for i in range(len(scraped_data['calls'])):
    row = scraped_data['general'][i] + scraped_data['strike'][i] + scraped_data['calls'][i] + scraped_data['puts'][i]
    data_list.append(row)

# Write to Excel and apply formatting
for r in data_list:
    worksheet.append(r)

# Define colors and alignment
green_fill = PatternFill(start_color="DFFFD6", end_color="DFFFD6", fill_type="solid")  # Light green
orange_fill = PatternFill(start_color="FFE4B2", end_color="FFE4B2", fill_type="solid")  # Light orange
purple_fill = PatternFill(start_color="E9D3FF", end_color="E9D3FF", fill_type="solid")  # Light purple
center_alignment = Alignment(horizontal="center", vertical="center")

# Apply formatting
for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row):
    for cell in row:
        if 'A' <= cell.column_letter <= 'F':
            cell.fill = green_fill
        elif 'G' <= cell.column_letter <= 'L':
            cell.fill = orange_fill
        elif 'M' <= cell.column_letter <= 'R':
            cell.fill = purple_fill
        cell.alignment = center_alignment

# Adjust column widths
for col in worksheet.columns:
    max_length = 0
    col_letter = col[0].column_letter  # Get column name
    for cell in col:
        if cell.value:
            max_length = max(max_length, len(str(cell.value)))
    worksheet.column_dimensions[col_letter].width = max_length + 2

# Save the Excel file
workbook.save(filename=file_name)
